import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
from openpyxl import Workbook
from openpyxl.drawing.image import Image

CSV_PATH = "bin/Debug/net8.0/sensor_data.csv"

# -------------------------------
# 📥 VERİ OKUMA FONKSİYONU
# -------------------------------
def load_data():
    try:
        df = pd.read_csv(CSV_PATH)

        df = df.tail(100)
        df.columns = ["Tarih", "Sicaklik", "Basinc", "Akim"]

        df["Tarih"] = pd.to_datetime(df["Tarih"], errors="coerce")

        df["Sicaklik"] = pd.to_numeric(df["Sicaklik"].astype(str).str.replace(",", "."), errors="coerce")
        df["Basinc"] = pd.to_numeric(df["Basinc"].astype(str).str.replace(",", "."), errors="coerce")
        df["Akim"] = pd.to_numeric(df["Akim"].astype(str).str.replace(",", "."), errors="coerce")

        df = df.dropna()
        df = df.sort_values("Tarih")

        return df

    except Exception as e:
        print("Veri okuma hatası:", e)
        return pd.DataFrame()

# -------------------------------
# 📊 İLK VERİ + EXCEL
# -------------------------------
df = load_data()

print("\n--- İSTATİSTİKLER ---")
print("Sicaklik Ortalama:", df["Sicaklik"].mean())
print("Basinc Ortalama:", df["Basinc"].mean())
print("Akim Ortalama:", df["Akim"].mean())

# Excel rapor
wb = Workbook()
ws = wb.active

ws["A1"] = "Sicaklik Ortalama"
ws["B1"] = df["Sicaklik"].mean()

ws["A2"] = "Basinc Ortalama"
ws["B2"] = df["Basinc"].mean()

ws["A3"] = "Akim Ortalama"
ws["B3"] = df["Akim"].mean()

wb.save("analysis_report.xlsx")

print("Excel rapor oluşturuldu!")

# -------------------------------
# 🚨 ANOMALİ
# -------------------------------
df["Anomali"] = df["Sicaklik"] > 50

print("\nAnomaliler:")
print(df[df["Anomali"] == True])

# -------------------------------
# 📈 REAL-TIME GRAFİK
# -------------------------------

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8))

def animate(i):
    df = load_data()

    if df.empty:
        return

    ax1.clear()
    ax2.clear()

    # 🔥 ÜST GRAFİK (3 VERİ)
    ax1.plot(df["Tarih"], df["Sicaklik"], color="red", label="Sicaklik")
    ax1.plot(df["Tarih"], df["Basinc"], color="blue", label="Basinc")
    ax1.plot(df["Tarih"], df["Akim"], color="green", label="Akim")

    ax1.set_title("Gercek Zamanli Veri")
    ax1.set_xlabel("Zaman")
    ax1.set_ylabel("Degerler")
    ax1.legend()
    ax1.grid(True)
    ax1.tick_params(axis='x', rotation=45)

    # 🔥 ALT GRAFİK (ANOMALİ)
    df["Anomali"] = (
    (df["Sicaklik"] > 50) |
    (df["Basinc"] > 9) |
    (df["Akim"] > 45)
)

    ax2.plot(df["Tarih"], df["Sicaklik"], color="red", label="Sicaklik")

    ax2.scatter(
        df[df["Anomali"]]["Tarih"],
        df[df["Anomali"]]["Sicaklik"],
        color="black",
        s=80,
        label="Anomali"
    )

    ax2.set_title("Anomali Analizi")
    ax2.set_xlabel("Zaman")
    ax2.set_ylabel("Sicaklik")
    ax2.legend()
    ax2.grid(True)
    ax2.tick_params(axis='x', rotation=45)

# ⚠️ EN KRİTİK SATIR
ani = FuncAnimation(fig, animate, interval=1000)

plt.tight_layout()
plt.show()